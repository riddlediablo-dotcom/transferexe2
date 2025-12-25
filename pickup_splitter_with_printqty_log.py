# -*- coding: utf-8 -*-
"""
工厂提货明细表自动拆分工具（带UI，保留模板公式，支持SKU/工厂配置）

新增/修复点（针对你这次的要求）：
1) 修复 Tkinter 异常弹窗的 NameError（Python 3 会清除 except e 变量）
2) 配置文件表头按你提供的格式：
   - sheet: SKU信息  (SKU / SKU检索 / 产品名称 / 工厂简称 / 箱规 / 长 / 宽 / 高 / 毛重 / 方数)
   - sheet: 工厂信息  (工厂名称 / 工厂地址)
   输出会把 SKU/长宽高/毛重 写入模板的「匹配」sheet（A~F固定结构）
3) 工厂地址支持“模糊匹配”：例如配置里工厂名称含“正美”，文件1供应商/工厂简称含“正美”也能匹配到地址
4) UI 增加勾选：是否按供应商建二级文件夹（勾上=拆分到供应商文件夹；不勾=直接输出到同一目录）
5) 输出路径自动建立子文件夹： 直发+YYYY.MM.DD（日期取 UI 的“预计提货日期”）
6) 模板文件：仍然建议选择（为了保留公式），但你只需要选一次（有记忆）。
   也支持把模板放在脚本/EXE同目录，命名为：工厂提货明细模板.xlsx（就可以不选）

依赖：
    pip install pandas openpyxl playwright

打包EXE（Windows）：
    pyinstaller --onefile --windowed pickup_splitter_ui_V4.py
"""

import os
import re
import json
import threading
import datetime
import time
import math
from uuid import uuid4
import copy
from typing import Callable, Optional, Dict, Any, Tuple, List

import pandas as pd
import numpy as np
import openpyxl
import requests
from openpyxl.formula.translate import Translator
from tkinter import ttk, filedialog, messagebox
import tkinter as tk
from tkinter.scrolledtext import ScrolledText


CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".pickup_splitter_config.json")

# ========= 积加 FBA 箱唛：查询→打印→传输中心下载 =========
BASE_URL = "https://gateway.apist.gerpgo.com"
DATA_GRID_URL = f"{BASE_URL}/supply/tms/query/shipment/dataGrid"
BATCH_PRINT_URL = f"{BASE_URL}/supply/tms/shipment/batchPrintLabels"
GET_DOWNLOAD_LIST_URL = f"{BASE_URL}/v2/download/reportDownload/getDownloadList"
GET_BATCH_FILE_URL = f"{BASE_URL}/v2/download/reportDownload/getBatchFileUrl"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36 Edg/143.0.0.0"
)

ZIP_PREFIX = "FBA_SHIPMENT_"
ZIP_SUFFIX = ".zip"

# --- FBA 打印限频（积加页面通常要求同一类“打印”操作间隔一段时间）---
FBA_PRINT_COOLDOWN_DEFAULT_SEC = 35
_FBA_LAST_PRINT_TS = 0.0
_FBA_PRINT_LOCK = threading.Lock()

def _fba_wait_cooldown(cooldown_sec: int, log_cb: Optional[Callable[[str], None]] = None):
    """确保两次 batchPrintLabels 提交之间至少间隔 cooldown_sec 秒。

    说明：积加前端通常会限制 30s 左右内重复点击“批量打印”，脚本太快会导致后续请求业务失败。
    """
    global _FBA_LAST_PRINT_TS
    try:
        cooldown_sec = int(float(cooldown_sec))
    except Exception:
        cooldown_sec = FBA_PRINT_COOLDOWN_DEFAULT_SEC
    if cooldown_sec < 0:
        cooldown_sec = 0

    with _FBA_PRINT_LOCK:
        now = time.time()
        wait = (_FBA_LAST_PRINT_TS + cooldown_sec) - now
        if wait <= 0:
            _FBA_LAST_PRINT_TS = now
            return

        end_ts = now + wait
        last_logged = None
        while True:
            remain = end_ts - time.time()
            if remain <= 0:
                _FBA_LAST_PRINT_TS = time.time()
                return
            sec = int(math.ceil(remain))

            # 日志不必每秒刷屏：每 10 秒提示一次；最后 3 秒每秒提示
            if log_cb:
                if sec <= 3 or (sec % 10 == 0):
                    if sec != last_logged:
                        last_logged = sec
                        log_cb(f"⏳ 等待 {sec}s（积加打印限频，默认 {cooldown_sec}s）…")

            time.sleep(1)

def _sanitize_header_value(v: str) -> str:
    if v is None:
        return ""
    if not isinstance(v, str):
        v = str(v)
    try:
        v.encode("latin-1")
        return v
    except UnicodeEncodeError:
        from urllib.parse import quote
        return quote(v, safe=":/;?&=,%+-_.~")

def _headers(token: str, cookie: str, page_url: str, page_title_encoded: str) -> Dict[str, str]:
    h = {
        "accept": "application/json, text/plain, */*",
        "accept-language": "zh-cn",
        "content-type": "application/json",
        "origin": "https://luteos.app.gerpgo.com",
        "referer": "https://luteos.app.gerpgo.com/",
        "user-agent": USER_AGENT,
        "x-auth-token": token,
        "x-api-id": str(uuid4()),
        "x-page-id": str(uuid4()),
        "x-page-title": page_title_encoded,
        "x-page-url": page_url,
        "Cookie": cookie,
    }
    return {k: _sanitize_header_value(v) for k, v in h.items()}

def _headers_fba(token: str, cookie: str) -> Dict[str, str]:
    return _headers(token, cookie, "/amzv-app/tms/fbaShipment", "FBA%E8%B4%A7%E4%BB%B6")

def _headers_tc(token: str, cookie: str) -> Dict[str, str]:
    return _headers(token, cookie, "/amzv-app/platform/reports/transmission-center", "%E4%BC%A0%E8%BE%93%E4%B8%AD%E5%BF%83")

def _request_json(session: requests.Session, method: str, url: str, *, headers: Dict[str, str], params=None, json_body=None, timeout=30) -> Tuple[int, Any, str]:
    resp = session.request(method, url, headers=headers, params=params, json=json_body, timeout=timeout)
    text = resp.text
    try:
        j = resp.json()
    except Exception:
        j = {"_raw_text": text}
    return resp.status_code, j, text

def _extract_grid_rows(grid_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    data = grid_json.get("data") or {}
    if isinstance(data, dict):
        for k in ["rows", "list", "records", "data", "result", "items"]:
            v = data.get(k)
            if isinstance(v, list) and (not v or isinstance(v[0], dict)):
                return v
    if isinstance(grid_json.get("data"), list):
        return grid_json["data"]
    return []

def _extract_download_rows(resp_json: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if isinstance(resp_json, dict):
        data = resp_json.get("data")
        if isinstance(data, dict):
            for k in ["list", "records", "rows", "data", "result", "items"]:
                v = data.get(k)
                if isinstance(v, list):
                    rows = v
                    break
    out = []
    for r in rows or []:
        if isinstance(r, dict) and (r.get("fileName") or r.get("filename")):
            out.append(r)
    return out

def _is_target_zip(row: Dict[str, Any]) -> bool:
    fn = (row.get("fileName") or row.get("filename") or "")
    return isinstance(fn, str) and fn.startswith(ZIP_PREFIX) and fn.lower().endswith(ZIP_SUFFIX)

def _parse_row_time(row: Dict[str, Any]) -> Optional[datetime.datetime]:
    t = row.get("requestTime") or row.get("gmtCreate") or row.get("createTime") or row.get("applyTime") or ""
    if isinstance(t, (int, float)):
        ts = t / 1000 if t > 10_000_000_000 else float(t)
        return datetime.datetime.fromtimestamp(ts)
    if isinstance(t, str) and t:
        try:
            return datetime.datetime.strptime(t[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None
    return None

def read_fba_ids_from_split_xlsx(xlsx_path: str) -> List[str]:
    """从拆分后的 Excel（sheet: 工厂提货明细）读取用于请求打印的 shipmentId 列表。

    约定：
    - 拆分文件里：Reference ID 列存放“FBA货件编号”（值通常包含/以 FBA 开头）
    - 仅对包含 'FBA' 的 Reference ID 发起打印/下载请求（TF 调拨单不参与）
    """
    try:
        df = pd.read_excel(xlsx_path, sheet_name="工厂提货明细", engine="openpyxl", dtype=str)
    except Exception:
        df = pd.read_excel(xlsx_path, sheet_name=0, engine="openpyxl", dtype=str)

    cols = [str(c).strip() for c in df.columns]
    cand = None
    for c in cols:
        if c in ("Reference ID", "Reference_ID", "reference_id", "参考单号", "ReferenceId"):
            cand = c
            break
    if not cand:
        # 兜底：包含 reference 的列
        for c in cols:
            if "reference" in c.lower():
                cand = c
                break
    if not cand:
        return []

    ids = []
    for v in df[cand].fillna("").astype(str).tolist():
        s = v.strip()
        if not s:
            continue
        su = s.upper()
        if "FBA" in su:
            ids.append(su)

    seen = set()
    out = []
    for x in ids:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out

def fba_download_labels_for_file(xlsx_path: str, token: str, cookie: str, log_cb: Optional[Callable[[str], None]]=None,
                                 poll_interval_sec: int = 3, poll_timeout_sec: int = 240, lookback_sec: int = 180,
                                 cooldown_sec: int = FBA_PRINT_COOLDOWN_DEFAULT_SEC) -> Optional[str]:
    """单文件：读取FBA ID → 查询→打印→轮询传输中心→下载ZIP到同目录。

    主要修复点：
    - 优先尝试使用 openpyxl.load_workbook(..., data_only=True) 来读取 sheet 的“计算后”值（如果单元格是公式且 Excel 已保存计算过的值，会返回计算值）。
    - 如果 data_only 读取后发货箱数仍为空，则尝试用 发货数量 / 单箱数量 向上取整 计算箱数。
    - 若仍无法获得，则回退使用 API 返回的 cartonQuantity/boxNum/packingBoxNum，最终回退 1。
    """
    # 先尝试读取拆分表（优先 sheet 名称包含 "工厂提货明细"）
    df = None
    workbook = None
    sheet_name_used = None
    try:
        # 1) 尝试 openpyxl data_only 方式读取（优先）
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
            # 选 sheet：优先名为 "工厂提货明细"，否则第一个
            sn = None
            for s in wb.sheetnames:
                if "工厂" in s or "提货" in s or "明细" in s:
                    sn = s
                    break
            if not sn:
                sn = wb.sheetnames[0]
            sheet_name_used = sn
            ws = wb[sn]
            # 将 sheet 内容转为 DataFrame
            data = []
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(x).strip() if x is not None else "" for x in row]
                    continue
                # 保证长度一致
                rvals = []
                for j in range(len(headers)):
                    if j < len(row):
                        rvals.append(row[j])
                    else:
                        rvals.append(None)
                data.append(rvals)
            if headers:
                df = pd.DataFrame(data, columns=headers, dtype=str)
                workbook = wb
        except Exception:
            # 忽略失败，后续尝试 pandas 读取
            df = None
            workbook = None

        # 2) 如果 openpyxl data_only 未得到有效 DataFrame，再用 pandas 读取（通常能读取但 formula 可能是公式文本）
        if df is None:
            try:
                try:
                    df = pd.read_excel(xlsx_path, sheet_name="工厂提货明细", engine="openpyxl", dtype=str)
                    sheet_name_used = "工厂提货明细"
                except Exception:
                    df = pd.read_excel(xlsx_path, sheet_name=0, engine="openpyxl", dtype=str)
                    sheet_name_used = df.columns.name if df.columns.name else 0
            except Exception:
                if log_cb:
                    log_cb(f"⚠️ 无法读取拆分文件以获取发货箱数，稍后将回退到 API 返回的箱数或默认 1：{xlsx_path}")
                df = None
    except Exception:
        df = None

    # 建立 Reference ID -> 发货箱数 映射
    id_to_qty = {}
    if df is not None and not df.empty:
        # 标准化列名列表
        cols = [str(c).strip() for c in df.columns]
        # 候选列 (优先级)
        id_col_candidates = ["Reference ID", "Reference_ID", "reference_id", "ReferenceId", "FBA ID", "FBA货件编号", "FBA货件号", "参考单号"]
        qty_col_candidates = ["发货箱数", "发货箱", "发货箱数量", "箱数", "箱数(发货箱数)", "发货箱数(J)"]
        # 查找 id 列（按候选优先级）
        id_col = None
        for cand in id_col_candidates:
            if cand in cols:
                id_col = cand
                break
        if not id_col:
            # 宽松匹配
            for c in cols:
                if "reference" in c.lower() or "fba" in c.lower() or "货件" in c:
                    id_col = c
                    break
        # 查找 qty 列
        qty_col = None
        for cand in qty_col_candidates:
            if cand in cols:
                qty_col = cand
                break
        if not qty_col:
            for c in cols:
                lc = c.lower()
                if "箱" in lc and "箱规" not in lc and "箱数(" not in lc:
                    qty_col = c
                    break

        # 查找发货数量 / 单箱数量 用于计算
        ship_qty_cols = [c for c in cols if c in ("发货数量", "发货总数", "数量", "出货数量", "total_qty", "TotalQty")]
        single_box_cols = [c for c in cols if c in ("单箱数量", "箱规", "箱内数量", "单箱数", "units_per_carton", "units_per_box", "箱内数量(每箱)")]
        # 兜底 heuristic
        if not ship_qty_cols:
            for c in cols:
                if "发货数量" in c or "发货总" in c or c == "发货":
                    ship_qty_cols.append(c)
        if not single_box_cols:
            for c in cols:
                if "单箱" in c or "箱规" in c or "箱内" in c:
                    single_box_cols.append(c)

        # 如果 id_col 为空，无法建立映射
        if id_col:
            # 遍历行建立映射
            for _, rr in df.iterrows():
                raw_id = rr.get(id_col)
                if raw_id is None:
                    continue
                raw_id_s = str(raw_id).strip()
                if not raw_id_s or raw_id_s.lower() in ("nan", "none"):
                    continue
                key = raw_id_s.upper()

                qval = None
                # 1) 尝试直接从 qty_col 读取（openpyxl data_only 可会把公式的计算值放在这里）
                if qty_col and qty_col in df.columns:
                    try:
                        qraw = rr.get(qty_col)
                        if pd.notna(qraw) and str(qraw).strip() not in ("", "nan"):
                            qval = int(float(str(qraw).strip()))
                    except Exception:
                        qval = None

                # 2) 如果 qval 无效，尝试用 发货数量 / 单箱数量 计算（向上取整）
                if (qval is None or qval <= 0) and ship_qty_cols and single_box_cols:
                    computed = None
                    for sq in ship_qty_cols:
                        for sb in single_box_cols:
                            try:
                                s_val = rr.get(sq)
                                b_val = rr.get(sb)
                                if s_val in (None, "") or b_val in (None, ""):
                                    continue
                                s_num = float(str(s_val).strip())
                                b_num = float(str(b_val).strip())
                                if b_num == 0:
                                    continue
                                computed = math.ceil(s_num / b_num)
                                if computed > 0:
                                    qval = int(computed)
                                    break
                            except Exception:
                                continue
                        if qval is not None and qval > 0:
                            break

                # 3) 尝试附近列（如果某些情况下 qty 放在 id 左右）
                if (qval is None or qval <= 0):
                    try:
                        cols_list = cols
                        id_idx = cols_list.index(id_col)
                        for offset in (1, -1, 2, -2, 3):
                            idx = id_idx + offset
                            if 0 <= idx < len(cols_list):
                                cand = cols_list[idx]
                                try:
                                    qraw = rr.get(cand)
                                    if pd.notna(qraw) and str(qraw).strip() not in ("", "nan"):
                                        qval = int(float(str(qraw).strip()))
                                        break
                                except Exception:
                                    continue
                    except Exception:
                        pass

                # 兜底为 1
                if qval is None or (isinstance(qval, (int, float)) and qval <= 0):
                    qval = 1

                id_to_qty[key] = int(qval)

    # 读取 FBA IDs（保留原项目的 read_fba_ids_from_split_xlsx 函数）
    fba_ids = read_fba_ids_from_split_xlsx(xlsx_path)
    if not fba_ids:
        if log_cb:
            log_cb(f"ℹ️ 未发现FBA ID，跳过箱唛：{os.path.basename(xlsx_path)}")
        return None

    sess = requests.Session()
    payload = {"__inner_refresh": True, "sort": "id", "order": "descend", "shipmentIdList": fba_ids, "type": "FBA", "page": 1, "pagesize": 200}
    st, j, raw = _request_json(sess, "POST", DATA_GRID_URL, headers=_headers_fba(token, cookie), json_body=payload, timeout=30)
    if st < 200 or st >= 300:
        raise RuntimeError(f"FBA查询失败 HTTP={st}：{raw[:300]}")
    rows = _extract_grid_rows(j)

    wanted = set([x.upper() for x in fba_ids])
    tasks = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        sid = r.get("shipmentId") or r.get("shipmentID") or r.get("shipment_id") or r.get("shipmentNo")
        if sid is None:
            continue
        sid_key = str(sid).upper().strip()
        if sid_key not in wanted:
            # 兼容大小写差异
            if sid_key not in wanted:
                continue
        internal_id = r.get("id")

        # 优先使用拆分文件映射的发货箱数
        qty = None
        if id_to_qty and sid_key in id_to_qty:
            qty = id_to_qty[sid_key]
        else:
            # 回退使用 API 返回的字段
            qty = r.get("cartonQuantity") or r.get("boxNum") or r.get("packingBoxNum") or r.get("cartonNum") or r.get("carton_count") or 1
            try:
                qty = max(1, int(str(qty).strip()))
            except Exception:
                qty = 1

        base_task = {
            "printQuantity": qty,
            "pageType": "PackageLabel_Thermal_100_100",
            "printType": "Package",
            "hideShipFrom": False,
            "hideShipTo": False,
            "reorderFlag": False,
            "waterMarkFlag": False,
            "productNameFlag": False,
            "waterMarkTemplateId": "",
        }
        if log_cb:
            log_cb(f"🧾 FBA {sid_key} → 打印箱数 print qty = {qty}")

        if internal_id is not None:
            tasks.append({"id": internal_id, **base_task})
        else:
            tasks.append({"shipmentNo": sid, **base_task})

    if not tasks:
        raise RuntimeError("FBA查询有返回，但未匹配到可打印任务（请检查shipmentId是否存在/一致）")

    _fba_wait_cooldown(cooldown_sec, log_cb=log_cb)

    submit_time = datetime.datetime.now()
    st, _, raw2 = _request_json(sess, "POST", BATCH_PRINT_URL, headers=_headers_fba(token, cookie), json_body=tasks, timeout=60)
    if st not in (200, 203):
        raise RuntimeError(f"提交打印失败 HTTP={st}：{raw2[:300]}")
    if log_cb:
        log_cb(f"🖨️ 已提交FBA箱唛打印：{os.path.basename(xlsx_path)}（{len(tasks)} 个任务）")

    start_day = (submit_time - datetime.timedelta(days=1)).date()
    end_day = datetime.datetime.now().date()
    params = {"order":"", "page":1, "pagesize":50, "startDate": start_day.strftime("%Y-%m-%d"), "endDate": end_day.strftime("%Y-%m-%d"), "dateType": 1}

    st, base_json, rawb = _request_json(sess, "GET", GET_DOWNLOAD_LIST_URL, headers=_headers_tc(token, cookie), params=params, timeout=30)
    if st < 200 or st >= 300:
        raise RuntimeError(f"获取下载列表失败（基线） HTTP={st}: {rawb[:200]}")
    base_ids = {str(r.get("id")) for r in _extract_download_rows(base_json) if r.get("id") is not None}

    earliest = submit_time - datetime.timedelta(seconds=lookback_sec)
    deadline = time.time() + poll_timeout_sec
    picked = None

    while time.time() < deadline:
        st, cur_json, rawc = _request_json(sess, "GET", GET_DOWNLOAD_LIST_URL, headers=_headers_tc(token, cookie), params=params, timeout=30)
        if st < 200 or st >= 300:
            raise RuntimeError(f"获取下载列表失败 HTTP={st}: {rawc[:200]}")
        rows = _extract_download_rows(cur_json)
        candidates = []
        for r in rows:
            if r.get("id") is None:
                continue
            if str(r.get("id")) in base_ids:
                continue
            if not _is_target_zip(r):
                continue
            rt = _parse_row_time(r)
            if rt and rt < earliest:
                continue
            candidates.append(r)
        if candidates:
            candidates.sort(key=lambda x: int(x.get("id") or 0), reverse=True)
            picked = candidates[0]
            break
        time.sleep(max(1, poll_interval_sec))

    if not picked:
        raise TimeoutError("等待下载ZIP超时（传输中心未出现本次新增FBA_SHIPMENT_*.zip）")

    file_id = picked.get("id")
    file_name = picked.get("fileName") or picked.get("filename")
    st, j3, raw3 = _request_json(sess, "POST", GET_BATCH_FILE_URL, headers=_headers_tc(token, cookie),
                                json_body=[{"id": file_id, "fileName": file_name}], timeout=30)
    if st < 200 or st >= 300:
        raise RuntimeError(f"获取下载URL失败 HTTP={st}：{raw3[:300]}")
    dl_url = j3.get("data") if isinstance(j3, dict) else None
    if not dl_url:
        raise RuntimeError("下载URL返回为空")

    out_dir = os.path.dirname(os.path.abspath(xlsx_path))
    out_zip = os.path.join(out_dir, file_name)

    with sess.get(dl_url, headers={"user-agent": USER_AGENT}, stream=True, timeout=180) as r:
        r.raise_for_status()
        with open(out_zip, "wb") as f:
            for chunk in r.iter_content(chunk_size=1024*256):
                if chunk:
                    f.write(chunk)

    if log_cb:
        log_cb(f"✅ 箱唛ZIP下载完成：{out_zip}")
    return out_zip


def auto_login_get_token_cookie(account: str, password: str, log_cb: Optional[Callable[[str], None]]=None) -> Tuple[str, str]:
    """可选：Playwright 自动登录获取 token/cookie（若失败可回退手动粘贴）。"""
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        raise RuntimeError("未安装 playwright：该EXE需要内置 playwright（打包环境 requirements.txt 加 playwright 并用 PyInstaller --collect-all playwright 重新打包）")

    login_url = "https://luteos.app.gerpgo.com/"
    token = ""
    cookie = ""

    with sync_playwright() as p:
        # 优先使用系统已安装的 Chrome（避免要求 playwright install 下载浏览器）
        try:
            browser = p.chromium.launch(channel="chrome", headless=False)
        except Exception:
            # 兜底：部分电脑可能没有 Chrome，但一定有 Edge
            browser = p.chromium.launch(channel="msedge", headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto(login_url, wait_until="domcontentloaded")
        page.wait_for_timeout(1200)

        user_selectors = [
            'input[placeholder*="账号"]',
            'input[placeholder*="用户名"]',
            'input[placeholder*="手机"]',
            'input[type="text"]',
        ]
        pwd_selectors = [
            'input[type="password"]',
            'input[placeholder*="密码"]',
        ]
        btn_selectors = [
            'button:has-text("登录")',
            'button:has-text("登 录")',
            'button[type="submit"]',
        ]

        def fill_first(selectors, value) -> bool:
            for sel in selectors:
                try:
                    el = page.query_selector(sel)
                    if el:
                        el.fill(value)
                        return True
                except Exception:
                    continue
            return False

        def click_first(selectors) -> bool:
            for sel in selectors:
                try:
                    el = page.query_selector(sel)
                    if el:
                        el.click()
                        return True
                except Exception:
                    continue
            return False

        if log_cb:
            log_cb("🌐 正在自动登录获取 token/cookie ...")
        if not fill_first(user_selectors, account) or not fill_first(pwd_selectors, password):
            browser.close()
            raise RuntimeError("自动登录失败：未找到账号/密码输入框（可能需要调整选择器）")
        if not click_first(btn_selectors):
            browser.close()
            raise RuntimeError("自动登录失败：未找到登录按钮（可能需要调整选择器）")

        page.wait_for_timeout(6000)

        try:
            token = page.evaluate("() => window.localStorage.getItem('x-auth-token') || window.localStorage.getItem('token') || ''")
        except Exception:
            token = ""

        ck = context.cookies()
        cookie = "; ".join([f"{c['name']}={c['value']}" for c in ck if c.get("name") and c.get("value")])
        browser.close()

    if not cookie:
        raise RuntimeError("自动登录未获取到 cookie")
    return token, cookie

DEFAULT_TEMPLATE_NAME = "工厂提货明细模板.xlsx"


# ========= 工具 =========
def sanitize_filename(s: str, replacement: str = "_") -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r'[\\/:*?"<>|\r\n]+', replacement, s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_date(date_str: str) -> Tuple[str, str]:
    """
    输入：2025-12-13 / 2025/12/13 / 2025.12.13 / 2025年12月13日
    输出：(写入单元格的格式 YYYY/MM/DD, 文件名格式 YYYY.MM.DD)
    """
    s = str(date_str).strip()
    s = s.replace("年", "-").replace("月", "-").replace("日", "")
    s = re.sub(r"[./]", "-", s)
    parts = [p for p in s.split("-") if p]
    if len(parts) < 3 or len(parts[0]) != 4:
        raise ValueError("日期格式请用 YYYY-MM-DD 或 YYYY/MM/DD 或 YYYY.MM.DD（例如 2025/12/13）")
    y, m, d = map(int, parts[:3])
    dt = datetime.date(y, m, d)
    return dt.strftime("%Y/%m/%d"), dt.strftime("%Y.%m.%d")



def export_mid_warehouse_keep_format(
    src_xlsx: str,
    sheet_name: str,
    header_row_1based: int,
    out_xlsx: str,
    log_cb=None,
):
    """导出“中仓/非工厂直发”的汇总文件，尽量保持与源文件一致的格式：
    - 保留源表头之前的所有行（含合并单元格、列宽、行高、样式）
    - 保留表头行样式
    - 仅保留满足条件的数据行（样式也一并复制）
    条件：
      1) “发运类型/直发类型”等列不包含“工厂直发”
      2) 渠道列包含 Amazon（amazon/亚马逊/amz）
    """
    wb = openpyxl.load_workbook(src_xlsx)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # 找表头行
    hr = header_row_1based
    # 读取表头
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hr, c).value
        headers.append(str(v).strip() if v is not None else "")

    def _find_col_idx(candidates):
        cands = [x.lower() for x in candidates]
        for i, h in enumerate(headers, start=1):
            hl = (h or "").strip().lower()
            if hl in cands:
                return i
        # 退而求其次：包含匹配（但仅用于非ID类字段）
        for i, h in enumerate(headers, start=1):
            hl = (h or "").strip().lower()
            for pat in cands:
                if pat and pat in hl:
                    return i
        return None

    direct_idx = _find_col_idx(["直发", "发运类型", "直发类型", "配送方式", "发货方式"])
    channel_idx = _find_col_idx(["渠道", "平台", "站点", "渠道名称", "渠道类型", "销售渠道"])

    # 如果缺列，就退化成原来的 df_non.to_excel（不阻断主流程）
    if direct_idx is None or channel_idx is None:
        if log_cb:
            log_cb("⚠️ 中仓汇总：未找到“发运类型/渠道”列，改用简化导出（可能不保留格式）。")
        df_tmp = pd.read_excel(src_xlsx, sheet_name=sheet_name, header=hr - 1, engine="openpyxl")
        df_tmp.to_excel(out_xlsx, index=False, engine="openpyxl")
        return

    # 新建输出工作簿
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title

    # 复制列宽
    for col_letter, dim in ws.column_dimensions.items():
        out_ws.column_dimensions[col_letter].width = dim.width

    # 复制合并单元格（先复制所有合并信息）
    for mr in ws.merged_cells.ranges:
        out_ws.merge_cells(str(mr))

    # 复制冻结窗格/筛选等（尽量）
    out_ws.freeze_panes = ws.freeze_panes
    if ws.auto_filter and ws.auto_filter.ref:
        out_ws.auto_filter.ref = ws.auto_filter.ref

    # 复制：表头之前行 + 表头行
    max_col = ws.max_column
    def _copy_cell(src_cell, dst_cell):
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell._style = copy.copy(src_cell._style)
        dst_cell.number_format = src_cell.number_format
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.comment = src_cell.comment

    for r in range(1, hr + 1):
        out_ws.row_dimensions[r].height = ws.row_dimensions[r].height
        for c in range(1, max_col + 1):
            _copy_cell(ws.cell(r, c), out_ws.cell(r, c))

    out_row = hr + 1

    # 过滤并复制数据行（样式保留）
    for r in range(hr + 1, ws.max_row + 1):
        direct_val = ws.cell(r, direct_idx).value
        direct_str = str(direct_val).replace(" ", "").replace("\u3000", "") if direct_val is not None else ""
        if "工厂直发" in direct_str:
            continue

        ch_val = ws.cell(r, channel_idx).value
        ch_str = str(ch_val).lower() if ch_val is not None else ""
        if not (("amazon" in ch_str) or ("亚马逊" in ch_str) or ("amz" in ch_str)):
            continue

        out_ws.row_dimensions[out_row].height = ws.row_dimensions[r].height
        for c in range(1, max_col + 1):
            _copy_cell(ws.cell(r, c), out_ws.cell(out_row, c))
        out_row += 1

    out_wb.save(out_xlsx)


def detect_sheet_and_header_row(xlsx_path: str) -> Tuple[str, int]:
    """
    自动在前50行里找包含“中仓”和“直发”的表头行，返回 (sheet_name, header_row_index_1based)
    找不到则默认第一个sheet第1行
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sh in wb.worksheets:
        maxrow = min(sh.max_row, 50)
        maxcol = min(sh.max_column, 80)
        for r in range(1, maxrow + 1):
            texts = []
            for c in range(1, maxcol + 1):
                v = sh.cell(r, c).value
                if isinstance(v, str):
                    texts.append(v.strip())
            if any(("中仓" in t and "直发" in t) for t in texts):
                return sh.title, r
    return wb.sheetnames[0], 1


def find_col(columns, candidates: List[str]) -> Optional[str]:
    cols = [str(c).strip() for c in columns]
    for pat in candidates:
        for c in cols:
            if c == pat:
                return c
    for pat in candidates:
        for c in cols:
            if pat in c:
                return c
    return None


def find_col_exact(columns, candidates: List[str]) -> Optional[str]:
    """只做精确列名匹配（避免把“发FBA数量”误当成ID列）。"""
    cols = [str(c).strip() for c in columns]
    for pat in candidates:
        for c in cols:
            if c == pat:
                return c
    return None


def choose_best_numeric_col(df: pd.DataFrame, base_name: str) -> Optional[str]:
    cand = [c for c in df.columns if str(c).strip() == base_name or str(c).startswith(base_name + ".")]
    best = None
    best_nonnull = -1
    for c in cand:
        s = pd.to_numeric(df[c], errors="coerce")
        nn = int(s.notna().sum())
        if nn > best_nonnull:
            best = c
            best_nonnull = nn
    return best


def supplier_short_name(s: str) -> str:
    if s is None:
        return "未知供应商"
    x = str(s).strip()
    x = re.sub(r'(有限责任公司|股份有限公司|有限公司|实业有限公司|实业|科技有限公司|科技|电器有限公司|电器|智能电器有限公司|智能|生物科技有限公司|生物科技|电子有限公司|电子|制造有限公司|制造|贸易有限公司|贸易)$', "", x)
    x = x.strip()
    # 去掉常见地域前缀（如：中山市/深圳市/广东省等），避免文件夹名过长
    x = re.sub(r'^(?:[\u4e00-\u9fff]{2,7}(?:省|市|自治区|自治州|地区|盟|州|县|区))', '', x)
    x = x.strip()
    # 尽量保留 2~6 个中文作为“短名”
    chs = re.findall(r'[\u4e00-\u9fff]+', x)
    if chs:
        t = chs[-1]
        if len(t) > 6:
            t = t[-6:]
        return sanitize_filename(t)
    return sanitize_filename(x[:10])


def norm_key(s: Any) -> str:
    if s is None:
        return ""
    x = str(s).strip()
    x = x.replace(" ", "").replace("\u3000", "")
    x = re.sub(r'[（）()【】\[\]{}<>《》“”"\'`·•,，.。:：;；\-_—/\\|]+', "", x)
    x = re.sub(r'(有限责任公司|股份有限公司|有限公司|实业有限公司|实业|科技有限公司|科技|电器有限公司|电器|智能电器有限公司|智能|生物科技有限公司|生物科技|电子有限公司|电子|制造有限公司|制造|贸易有限公司|贸易)$', "", x)
    return x

def norm_id_value(v: Any) -> str:
    """把单元格值规范为可用的字符串ID；None/NaN/空白都返回空串。"""
    if v is None:
        return ""
    try:
        # pandas 的 NaN / NaT
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return ""
    return s


def pick_first_id(*vals: Any) -> str:
    """按顺序取第一个非空ID（优先FBA，其次TF）。"""
    for v in vals:
        s = norm_id_value(v)
        if s:
            return s
    return ""




# ========= 读取配置（按你上传的表头） =========
def load_config_xlsx(cfg_path: str) -> Tuple[pd.DataFrame, Dict[str, str], Dict[str, str]]:
    """
    返回：
      sku_cfg_df：用于写入模板「匹配」sheet，字段至少含：
        SKU, 产品名称, 长, 宽, 高, 单箱毛重, 单箱数量
      sku_factory_short：SKU -> 工厂简称
      factory_name_to_addr：工厂名称 -> 工厂地址
    """
    xls = pd.read_excel(cfg_path, sheet_name=None, engine="openpyxl")

    if "SKU信息" not in xls:
        raise ValueError("配置文件缺少 sheet：SKU信息")
    df_sku = xls["SKU信息"].copy()
    df_sku.columns = [str(c).strip() for c in df_sku.columns]

    # 必要列（按你给的表头）
    sku_col = find_col(df_sku.columns, ["SKU"])
    sku_search_col = find_col(df_sku.columns, ["SKU检索"])
    name_col = find_col(df_sku.columns, ["产品名称"])
    fac_short_col = find_col(df_sku.columns, ["工厂简称"])
    carton_col = find_col(df_sku.columns, ["箱规"])
    l_col = find_col(df_sku.columns, ["长"])
    w_col = find_col(df_sku.columns, ["宽"])
    h_col = find_col(df_sku.columns, ["高"])
    gw_col = find_col(df_sku.columns, ["毛重"])

    if sku_col is None:
        raise ValueError("配置文件 SKU信息 缺少列：SKU")

    rows = []
    sku_factory_short: Dict[str, str] = {}

    for _, r in df_sku.iterrows():
        sku = str(r.get(sku_col)).strip() if r.get(sku_col) is not None else ""
        if not sku or sku.lower() in ("nan", "none"):
            continue

        row = {
            "SKU": sku,
            "产品名称": str(r.get(name_col)).strip() if name_col and r.get(name_col) is not None else "",
            "长": pd.to_numeric(r.get(l_col), errors="coerce") if l_col else np.nan,
            "宽": pd.to_numeric(r.get(w_col), errors="coerce") if w_col else np.nan,
            "高": pd.to_numeric(r.get(h_col), errors="coerce") if h_col else np.nan,
            "单箱毛重": pd.to_numeric(r.get(gw_col), errors="coerce") if gw_col else np.nan,
            "单箱数量": pd.to_numeric(r.get(carton_col), errors="coerce") if carton_col else np.nan,
        }
        rows.append(row)

        if fac_short_col:
            fs = str(r.get(fac_short_col)).strip() if r.get(fac_short_col) is not None else ""
            if fs and fs.lower() not in ("nan", "none"):
                sku_factory_short[sku] = fs

        # 兼容：如果 SKU检索 和 SKU 不同，也写一行“别名”，防止文件1用的是 SKU检索
        if sku_search_col:
            alias = str(r.get(sku_search_col)).strip() if r.get(sku_search_col) is not None else ""
            if alias and alias.lower() not in ("nan", "none") and alias != sku:
                alias_row = row.copy()
                alias_row["SKU"] = alias
                rows.append(alias_row)
                if fac_short_col and sku in sku_factory_short:
                    sku_factory_short[alias] = sku_factory_short[sku]

    sku_cfg_df = pd.DataFrame(rows, columns=["SKU", "产品名称", "长", "宽", "高", "单箱毛重", "单箱数量"]).drop_duplicates(subset=["SKU"], keep="last")

    # 工厂信息表
    factory_name_to_addr: Dict[str, str] = {}
    if "工厂信息" in xls:
        df_f = xls["工厂信息"].copy()
        df_f.columns = [str(c).strip() for c in df_f.columns]
        n_col = find_col(df_f.columns, ["工厂名称"])
        a_col = find_col(df_f.columns, ["工厂地址"])
        if n_col and a_col:
            for _, r in df_f.iterrows():
                n = str(r.get(n_col)).strip() if r.get(n_col) is not None else ""
                a = str(r.get(a_col)).strip() if r.get(a_col) is not None else ""
                if n and a and n.lower() not in ("nan", "none") and a.lower() not in ("nan", "none"):
                    factory_name_to_addr[n] = a

    return sku_cfg_df, sku_factory_short, factory_name_to_addr


def merge_missing_skus_from_file1(sku_cfg_df: pd.DataFrame, df1: pd.DataFrame) -> pd.DataFrame:
    sku_col = find_col(df1.columns, ["仓库SKU", "SKU"])
    name_col = find_col(df1.columns, ["产品名称", "品名"])
    if sku_col is None:
        return sku_cfg_df

    existing = set(sku_cfg_df["SKU"].astype(str).tolist()) if not sku_cfg_df.empty else set()
    add_rows = []
    for _, r in df1.iterrows():
        sku = str(r.get(sku_col)).strip() if r.get(sku_col) is not None else ""
        if not sku or sku.lower() in ("nan", "none"):
            continue
        if sku in existing:
            continue
        add_rows.append({
            "SKU": sku,
            "产品名称": str(r.get(name_col)).strip() if name_col and r.get(name_col) is not None else "",
            "长": np.nan,
            "宽": np.nan,
            "高": np.nan,
            "单箱毛重": np.nan,
            "单箱数量": np.nan,
        })
        existing.add(sku)

    if add_rows:
        sku_cfg_df = pd.concat([sku_cfg_df, pd.DataFrame(add_rows)], ignore_index=True)

    return sku_cfg_df


def fuzzy_factory_address(keys: List[str], factory_name_to_addr: Dict[str, str]) -> str:
    """
    keys：比如 [工厂简称, 供应商短名, 供应商全名]
    返回：匹配到的工厂地址（支持模糊包含）
    """
    if not factory_name_to_addr:
        return ""

    # 预处理
    fac_items = [(k, norm_key(k), v) for k, v in factory_name_to_addr.items()]
    best = ("", 0, "")  # name, score, addr

    for key in keys:
        nk = norm_key(key)
        if not nk or len(nk) < 2:
            continue
        for orig_name, nn, addr in fac_items:
            score = 0
            if nk in nn:
                score = len(nk)
            elif nn in nk:
                score = len(nn)
            if score > best[1]:
                best = (orig_name, score, addr)

    return best[2] if best[1] > 0 else ""


def fuzzy_factory_name(keys: List[str], factory_name_to_addr: Dict[str, str]) -> str:
    """
    keys：比如 [工厂简称, 供应商短名, 供应商全名]
    返回：匹配到的“工厂名称”（配置表里的名字，支持模糊包含）
    """
    if not factory_name_to_addr:
        return ""

    fac_items = [(k, norm_key(k), v) for k, v in factory_name_to_addr.items()]
    best = ("", 0)  # name, score

    for key in keys:
        nk = norm_key(key)
        if not nk or len(nk) < 2:
            continue

        for orig_name, nn, _addr in fac_items:
            score = 0
            if nk == nn:
                score = 1000 + len(nk)
            elif nk in nn:
                score = len(nk)
            elif nn in nk:
                score = len(nn)
            if score > best[1]:
                best = (orig_name, score)

    return best[0] if best[1] > 0 else ""



# ========= 模板写入（保留公式） =========
def _copy_cell_style(src, dst):
    """安全复制样式：避免 StyleProxy 导致的 'unhashable type: StyleProxy'"""
    try:
        if getattr(src, "has_style", False):
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    except Exception:
        # 样式复制失败不影响数据/公式输出
        pass



def write_match_sheet(wb, sku_cfg_df: pd.DataFrame):
    """
    把配置写入模板的「匹配」sheet，列顺序必须是：
      A SKU, B 产品名称, C 长, D 宽, E 高, F 单箱毛重, G 单箱数量(可有可无，但保留)
    这样模板里的：VLOOKUP($E2,匹配!$A:$F,3..6) 才能正常取到长宽高/毛重
    """
    if "匹配" not in wb.sheetnames:
        wb.create_sheet("匹配")
    ws = wb["匹配"]

    ws.delete_rows(1, ws.max_row if ws.max_row > 0 else 1)

    headers = ["SKU", "产品名称", "长", "宽", "高", "单箱毛重", "单箱数量"]
    ws.append(headers)

    if sku_cfg_df is None or sku_cfg_df.empty:
        return

    for _, r in sku_cfg_df.iterrows():
        ws.append([
            str(r.get("SKU")).strip() if r.get("SKU") is not None else "",
            str(r.get("产品名称")).strip() if r.get("产品名称") is not None else "",
            None if pd.isna(r.get("长")) else float(r.get("长")),
            None if pd.isna(r.get("宽")) else float(r.get("宽")),
            None if pd.isna(r.get("高")) else float(r.get("高")),
            None if pd.isna(r.get("单箱毛重")) else float(r.get("单箱毛重")),
            None if pd.isna(r.get("单箱数量")) else float(r.get("单箱数量")),
        ])


def rebuild_main_sheet_with_data(
    ws,
    data_rows: List[Dict[str, Any]],
    pickup_date_cell: str,
    template_data_row: int = 2,
    template_total_row: int = 4,
):
    """
    用模板第2行作为“数据行模板”，模板第4行作为“合计行模板”
    - 克隆样式+公式到 N 行
    - 用表头匹配写入需要写值的列（不覆盖公式列）
    """
    max_col = ws.max_column

    tmpl_cells = [ws.cell(template_data_row, c) for c in range(1, max_col + 1)]
    total_cells = [ws.cell(template_total_row, c) for c in range(1, max_col + 1)]

    tmpl_height = ws.row_dimensions[template_data_row].height
    total_height = ws.row_dimensions[template_total_row].height

    # 删除旧数据区域
    last = ws.max_row
    if last >= template_data_row:
        ws.delete_rows(template_data_row, last - template_data_row + 1)

    start_row = template_data_row
    n = len(data_rows)
    if n <= 0:
        return

    # 插入 n 数据行 + 1 合计行
    ws.insert_rows(start_row, amount=n + 1)

    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    col_map = {str(h).strip(): idx for idx, h in enumerate(headers, start=1) if h is not None}

    def setv(row_idx: int, col_name: str, val: Any):
        if col_name in col_map:
            ws.cell(row_idx, col_map[col_name]).value = val

    # 写数据行
    for i, row_data in enumerate(data_rows):
        r = start_row + i

        # 克隆模板行（样式+公式）
        for c in range(1, max_col + 1):
            src = tmpl_cells[c - 1]
            dst = ws.cell(r, c)
            _copy_cell_style(src, dst)

            if isinstance(src.value, str) and src.value.startswith("="):
                dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
            else:
                # 不写死值，后面用 setv 写入需要写值的列；其余保持空/由公式列负责
                dst.value = None

        if tmpl_height is not None:
            ws.row_dimensions[r].height = tmpl_height

        # 写值列（按表头名）
        setv(r, "预计提货日期", pickup_date_cell)
        setv(r, "销售负责人", row_data.get("销售负责人"))
        setv(r, "账号", row_data.get("账号"))
        setv(r, "FNSKU / UPC", row_data.get("FNSKU / UPC"))
        setv(r, "SKU", row_data.get("SKU"))
        setv(r, "产品名称", row_data.get("产品名称"))
        setv(r, "发货数量", row_data.get("发货数量"))
        setv(r, "单箱数量", row_data.get("单箱数量"))
        setv(r, "物流渠道", row_data.get("物流渠道"))
        setv(r, "发货仓库", row_data.get("发货仓库"))
        setv(r, "FBA ID", row_data.get("FBA ID"))
        setv(r, "Reference ID", row_data.get("Reference ID"))
        setv(r, "到货仓库", row_data.get("到货仓库"))
        setv(r, "仓库代码", row_data.get("仓库代码"))
        setv(r, "工厂地址", row_data.get("工厂地址"))

    # 合计行
    total_row = start_row + n
    for c in range(1, max_col + 1):
        src = total_cells[c - 1]
        dst = ws.cell(total_row, c)
        _copy_cell_style(src, dst)
        dst.value = None

    if total_height is not None:
        ws.row_dimensions[total_row].height = total_height

    # 重新写合计列（如果模板就是 SUM 也行，这里强制按实际范围）
    def set_sum(col_letter: str):
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws.cell(total_row, col_idx).value = f"=SUM({col_letter}{start_row}:{col_letter}{start_row + n - 1})"

    for col_letter in ["G", "J", "V", "W", "X"]:
        try:
            set_sum(col_letter)
        except Exception:
            pass

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ========= 主流程 =========
def build_data_rows_from_file1(
    df: pd.DataFrame,
    sku_cfg_df: pd.DataFrame,
    sku_factory_short: Dict[str, str],
    factory_name_to_addr: Dict[str, str],
    supplier_value: Any,
) -> List[Dict[str, Any]]:
    """
    把文件1的行映射为模板需要的“值列”
    - 销售负责人：来自文件1列“运营”
    - 工厂地址：优先用 SKU->工厂简称 -> 工厂信息模糊匹配；再用供应商名模糊匹配
    - 单箱数量：优先文件1的“箱规”，否则用配置表 sku_cfg_df 的单箱数量
    """
    op_col = find_col(df.columns, ["运营"])
    acct_col = find_col(df.columns, ["店铺账号/目的仓库", "账号"])
    fns_col = find_col(df.columns, ["FNSKU / UPC", "FNSKU/UPC", "FNSKU"])
    sku_col = find_col(df.columns, ["仓库SKU", "SKU"])
    prod_col = find_col(df.columns, ["产品名称", "品名"])
    qty_col = find_col(df.columns, ["发货数量", "数量"])
    carton_col = choose_best_numeric_col(df, "箱规")
    ship_mode_col = find_col(df.columns, ["物流渠道", "物流方式"])
    ship_from_col = find_col(df.columns, ["发货仓库", "发货仓"])
    fba_col = find_col_exact(df.columns, ["FBA货件编号", "FBA ID", "FBA货件ID", "FBA货件号"])
    ref_col = find_col_exact(df.columns, ["TF调拨单", "TF调拨单号", "调拨单号", "TF单号", "调拨单", "Reference ID", "参考单号"])
    dest_col = find_col(df.columns, ["配送地址/收货人信息", "到货仓库"])
    wh_code_col = find_col(df.columns, ["仓库代码"])

    cfg_carton_map = {}
    if sku_cfg_df is not None and not sku_cfg_df.empty:
        # SKU -> 单箱数量
        tmp = sku_cfg_df[["SKU", "单箱数量"]].copy()
        tmp["SKU"] = tmp["SKU"].astype(str)
        cfg_carton_map = dict(zip(tmp["SKU"], tmp["单箱数量"]))

    supplier_full = str(supplier_value).strip() if supplier_value is not None else ""
    supplier_short = supplier_short_name(supplier_full)

    rows: List[Dict[str, Any]] = []
    for _, r in df.iterrows():
        sku = str(r.get(sku_col)).strip() if sku_col and r.get(sku_col) is not None else ""
        fac_short = sku_factory_short.get(sku, "")
        addr = fuzzy_factory_address([fac_short, supplier_short, supplier_full], factory_name_to_addr)

        # 单箱数量（箱规）
        carton = None
        if carton_col:
            carton = pd.to_numeric(r.get(carton_col), errors="coerce")
            carton = None if pd.isna(carton) else float(carton)
        if carton is None and sku in cfg_carton_map:
            v = cfg_carton_map.get(sku)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                carton = float(v)

        rows.append({
            "销售负责人": r.get(op_col) if op_col else None,
            "账号": r.get(acct_col) if acct_col else None,
            "FNSKU / UPC": r.get(fns_col) if fns_col else None,
            "SKU": sku if sku else None,
            "产品名称": r.get(prod_col) if prod_col else None,
            "发货数量": r.get(qty_col) if qty_col else None,
            "单箱数量": carton,
            "物流渠道": r.get(ship_mode_col) if ship_mode_col else None,
            "发货仓库": r.get(ship_from_col) if ship_from_col else None,
            "FBA ID": (norm_id_value(r.get(ref_col)) if ref_col else ""),
            "Reference ID": (norm_id_value(r.get(fba_col)) if fba_col else ""),
            "到货仓库": r.get(dest_col) if dest_col else None,
            "仓库代码": r.get(wh_code_col) if wh_code_col else None,
            "工厂地址": addr,
        })
    return rows



def choose_shipment_folder_id(df: pd.DataFrame, fba_col: Optional[str], ref_col: Optional[str]) -> str:
    """
    拆分后输出文件统一放到“FBA ID”文件夹逻辑：
    - 优先用 FBA货件编号/FBA ID（即使是中仓也一样）
    - 如果没有FBA，则使用TF调拨单/调拨单号（仍然归入同一层ID文件夹，不再单独建TF目录）
    """
    def _pick_first(colname: Optional[str]) -> str:
        if not colname or colname not in df.columns:
            return ""
        for v in df[colname].fillna("").astype(str).tolist():
            s = v.strip()
            if s:
                return s
        return ""

    sid = _pick_first(fba_col)
    if not sid:
        sid = _pick_first(ref_col)

    sid = (sid or "UNKNOWN").strip()
    # folder name safe
    sid = sanitize_filename(sid.upper())
    return sid


def resolve_template_path(template_input: str) -> str:
    t = (template_input or "").strip()
    if t and os.path.isfile(t):
        return t

    # 如果用户没选模板，就尝试脚本/EXE同目录的默认模板名
    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidate = os.path.join(base_dir, DEFAULT_TEMPLATE_NAME)
    if os.path.isfile(candidate):
        return candidate

    raise ValueError("找不到模板文件。请在UI里选择“文件2模板（含公式）”，或把模板放到程序同目录并命名为：工厂提货明细模板.xlsx")


def process_file(
    file1: str,
    template_path_input: str,
    cfg_path: str,
    out_root: str,
    pickup_date: str,
    time_tag: str,
    product_tag: str,
    filename_name: str,
    split_supplier_folder: bool,
    progress_cb: Optional[Callable[[int, int, str], None]] = None,
    log_cb: Optional[Callable[[str], None]] = None,
) -> List[str]:
    pickup_cell, pickup_fname = parse_date(pickup_date)

    # 输出：自动创建 直发MMDD 文件夹
    out_base = os.path.join(out_root, f"直发{pickup_fname[5:7]}{pickup_fname[8:10]}")
    os.makedirs(out_base, exist_ok=True)

    # 读文件1
    sheet, header_row = detect_sheet_and_header_row(file1)
    df1 = pd.read_excel(file1, sheet_name=sheet, header=header_row - 1, engine="openpyxl")
    df1.columns = [str(c).strip() if c is not None else "" for c in df1.columns]

    direct_col = find_col(df1.columns, ["中仓 或 工厂直发", "中仓或工厂直发", "工厂直发"])
    supplier_col = find_col(df1.columns, ["供应商", "供应商名称", "工厂"])
    if direct_col is None or supplier_col is None:
        raise ValueError(f"找不到必要列：{direct_col=} , {supplier_col=}。请确认文件1表头是否一致。")

    ser = df1[direct_col].astype(str).str.replace(" ", "").str.replace("\u3000", "")
    df_f = df1[ser.str.contains("工厂直发", na=False)].copy()
    if df_f.empty:
        return []

    # 非“工厂直发”的行（例如：中仓）——用于生成一个汇总总表
    df_non = df1[~ser.str.contains("工厂直发", na=False)].copy()

    # 仅保留 Amazon（中仓汇总只要亚马逊，过滤掉 Shopify/Walmart 等）
    try:
        ch_col = find_col(df_non.columns, ["渠道", "平台", "平台站点", "站点", "Channel", "Platform", "店铺", "账号", "账户"])
        if ch_col is not None:
            _ser = df_non[ch_col].astype(str).str.lower()
            df_non = df_non[_ser.str.contains("amazon|亚马逊|amz", na=False)].copy()
    except Exception:
        pass


    # 读配置
    sku_cfg_df, sku_factory_short, factory_name_to_addr = load_config_xlsx(cfg_path)
    sku_cfg_df = merge_missing_skus_from_file1(sku_cfg_df, df_f)

    # 模板路径
    template_path = resolve_template_path(template_path_input)

    outputs = []
    # 统一归档ID：优先 FBA货件编号/FBA ID；若无则用 TF调拨单/调拨单号
    fba_col = find_col_exact(df_f.columns, ["FBA货件编号", "FBA ID", "FBA货件ID", "FBA货件号"])
    ref_col = find_col_exact(df_f.columns, ["TF调拨单", "TF调拨单号", "调拨单号", "TF单号", "调拨单", "Reference ID", "参考单号"])

    groups = list(df_f.groupby(supplier_col))
    total = len(groups)

    # 先预计算：每个供应商对应的“标准工厂文件夹名”
    supplier_to_factory: Dict[str, str] = {}
    factory_to_suppliers: Dict[str, set] = {}
    for supplier, _g in groups:
        sup_short = supplier_short_name(supplier)
        factory_folder = fuzzy_factory_name([sup_short, str(supplier)], factory_name_to_addr) or sup_short
        supplier_to_factory[str(supplier)] = factory_folder
        factory_to_suppliers.setdefault(factory_folder, set()).add(sup_short)

    for i, (supplier, g) in enumerate(groups, start=1):
        sup_short = supplier_short_name(supplier)

        # 工厂文件夹名：使用预计算结果（配置表匹配优先）
        factory_folder = supplier_to_factory.get(str(supplier), sup_short)

        # 输出目录：
        # - 只拆分到“工厂”（不再创建 FBA/TF 的子文件夹）
        # - 不勾：输出/直发MMDD/...（全部文件直接放在直发MMDD根目录）
        # - 勾上：输出/直发MMDD/工厂(配置名)/...（只到工厂这一层，不再创建供应商第三级）
        if split_supplier_folder:
            folder = os.path.join(out_base, factory_folder)
        else:
            folder = out_base


        os.makedirs(folder, exist_ok=True)

        # 数据行映射
        data_rows = build_data_rows_from_file1(
            df=g,
            sku_cfg_df=sku_cfg_df,
            sku_factory_short=sku_factory_short,
            factory_name_to_addr=factory_name_to_addr,
            supplier_value=supplier,
        )

        # 打开模板 + 写匹配sheet + 写主表
        wb = openpyxl.load_workbook(template_path)
                # 仅保留本次拆分涉及的 SKU（减少匹配表冗余）
        try:
            sku_in_file = {str(rr.get('SKU')).strip() for rr in data_rows if rr.get('SKU') is not None}
            sku_in_file = {s for s in sku_in_file if s}
            sku_cfg_sub = sku_cfg_df[sku_cfg_df['SKU'].astype(str).str.strip().isin(sku_in_file)].copy() if (sku_cfg_df is not None and not sku_cfg_df.empty and sku_in_file) else sku_cfg_df
        except Exception:
            sku_cfg_sub = sku_cfg_df
        write_match_sheet(wb, sku_cfg_sub)

        if "工厂提货明细" not in wb.sheetnames:
            raise ValueError("模板文件缺少 sheet：工厂提货明细")
        ws = wb["工厂提货明细"]

        rebuild_main_sheet_with_data(
            ws=ws,
            data_rows=data_rows,
            pickup_date_cell=pickup_cell,
            template_data_row=2,
            template_total_row=4,
        )

        # 保存
        # 文件名：姓名-【日期+时间(可选) + 产品(可选)+供应商】工厂提货明细表
        tag = pickup_fname
        t = sanitize_filename(time_tag) if time_tag else ""
        p = sanitize_filename(product_tag) if product_tag else ""
        if t:
            tag += f"+{t}"
        if p:
            sep = " + " if t else "+"  # 只有时间存在时，用 “ + ” 分隔产品，符合你示例
            tag += f"{sep}{p}"
        tag += f"+{sup_short}"
        filename = f"{sanitize_filename(filename_name)}-【{tag}】工厂提货明细表.xlsx"
        out_path = os.path.join(folder, filename)

        base, ext = os.path.splitext(out_path)
        k = 1
        while os.path.exists(out_path):
            out_path = f"{base}({k}){ext}"
            k += 1

        wb.save(out_path)
        outputs.append(out_path)

        if log_cb:
            log_cb(f"✅ {sup_short} -> {out_path}")

        if progress_cb:
            progress_cb(i, total, sup_short)
    # 生成“中仓直发YYYYMMDD”汇总表（放在 out_base 目录）
    if df_non is not None and not df_non.empty:
        try:
            # 数据行：按供应商分组生成（确保工厂地址模糊匹配按供应商生效）
            data_rows_non = []
            if supplier_col is not None:
                for sup, gg in df_non.groupby(supplier_col):
                    data_rows_non.extend(build_data_rows_from_file1(
                        df=gg,
                        sku_cfg_df=sku_cfg_df,
                        sku_factory_short=sku_factory_short,
                        factory_name_to_addr=factory_name_to_addr,
                        supplier_value=sup,
                    ))
            else:
                data_rows_non = build_data_rows_from_file1(
                    df=df_non,
                    sku_cfg_df=sku_cfg_df,
                    sku_factory_short=sku_factory_short,
                    factory_name_to_addr=factory_name_to_addr,
                    supplier_value="",
                )
            # 中仓/非工厂直发：保持原始格式输出（不套模板）
            

            yyyymmdd = pickup_fname.replace(".", "")
            sum_name = f"中仓{yyyymmdd}.xlsx"
            sum_path = os.path.join(out_base, sum_name)

            base, ext = os.path.splitext(sum_path)
            k = 1
            while os.path.exists(sum_path):
                sum_path = f"{base}({k}){ext}"
                k += 1

            export_mid_warehouse_keep_format(file1, sheet, header_row, sum_path, log_cb=log_cb)

            if log_cb:
                log_cb(f"📌 汇总表 -> {sum_path}")
        except Exception as _ex:
            # 汇总表失败不影响主流程
            if log_cb:
                log_cb(f"⚠️ 汇总表生成失败：{_ex}")


    return outputs


# ========= UI =========
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("工厂提货明细表自动拆分（保留模板公式 + SKU/工厂配置）")
        self.geometry("980x640")
        self.resizable(True, True)

        self.file1_var = tk.StringVar()
        self.template_var = tk.StringVar()   # 可选：不填则找同目录默认模板
        self.cfg_var = tk.StringVar()
        self.outdir_var = tk.StringVar()
        self.date_var = tk.StringVar()
        self.time_var = tk.StringVar()
        self.product_var = tk.StringVar()
        self.name_var = tk.StringVar()
        self.split_var = tk.BooleanVar(value=True)
        # --- 积加认证（用于下载 FBA 箱唛）---
        self.token_var = tk.StringVar()
        self.acc_var = tk.StringVar()
        self.pwd_var = tk.StringVar()
        self.enable_fba_label_var = tk.BooleanVar(value=True)
        self.fba_cooldown_var = tk.StringVar(value=str(FBA_PRINT_COOLDOWN_DEFAULT_SEC))

        self._build_ui()
        self._load_config()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_ui(self):
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)
        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(12, weight=1)

        r = 0
        ttk.Label(frm, text="文件1（线上取回数据）").grid(row=r, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.file1_var).grid(row=r, column=1, sticky="we", padx=8)
        ttk.Button(frm, text="浏览…", command=self._pick_file1).grid(row=r, column=2, sticky="e")

        r += 1
        ttk.Label(frm, text=f"文件2模板（含公式，可选；不选则自动找同目录：{DEFAULT_TEMPLATE_NAME}）").grid(row=r, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=self.template_var).grid(row=r, column=1, sticky="we", padx=8, pady=(8, 0))
        ttk.Button(frm, text="选择…", command=self._pick_template).grid(row=r, column=2, sticky="e", pady=(8, 0))

        r += 1
        ttk.Label(frm, text="配置文件（你给的SKU信息/工厂信息）").grid(row=r, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=self.cfg_var).grid(row=r, column=1, sticky="we", padx=8, pady=(8, 0))
        ttk.Button(frm, text="选择…", command=self._pick_cfg).grid(row=r, column=2, sticky="e", pady=(8, 0))

        r += 1
        ttk.Label(frm, text="输出根目录（程序会自动在里面建：直发MMDD，例如 直发1225）").grid(row=r, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=self.outdir_var).grid(row=r, column=1, sticky="we", padx=8, pady=(8, 0))
        ttk.Button(frm, text="选择…", command=self._pick_outdir).grid(row=r, column=2, sticky="e", pady=(8, 0))

        r += 1
        sub = ttk.Frame(frm)
        sub.grid(row=r, column=0, columnspan=3, sticky="we", pady=(10, 0))
        # 让中间输入框可扩展
        sub.columnconfigure(1, weight=0)
        sub.columnconfigure(3, weight=0)
        sub.columnconfigure(5, weight=0)
        sub.columnconfigure(7, weight=1)

        ttk.Label(sub, text="预计提货日期").grid(row=0, column=0, sticky="w")
        ttk.Entry(sub, textvariable=self.date_var, width=14).grid(row=0, column=1, sticky="w", padx=(8, 16))

        ttk.Label(sub, text="时间（选填，如：13点）").grid(row=0, column=2, sticky="w")
        ttk.Entry(sub, textvariable=self.time_var, width=10).grid(row=0, column=3, sticky="w", padx=(8, 16))

        ttk.Label(sub, text="产品（选填，如：空滤）").grid(row=0, column=4, sticky="w")
        ttk.Entry(sub, textvariable=self.product_var, width=12).grid(row=0, column=5, sticky="w", padx=(8, 16))

        ttk.Label(sub, text="姓名").grid(row=0, column=6, sticky="w")
        ttk.Entry(sub, textvariable=self.name_var, width=12).grid(row=0, column=7, sticky="w", padx=(8, 0))

        r += 1
        opt = ttk.Frame(frm)
        opt.grid(row=r, column=0, columnspan=3, sticky="we", pady=(10, 0))
        ttk.Checkbutton(opt, text="按供应商建立二级文件夹（勾上：输出/直发MMDD/供应商/…；不勾：输出/直发MMDD/…）", variable=self.split_var).pack(side="left")
        # --- FBA 箱唛下载（增量区域，不影响拆分逻辑） ---
        r += 1
        auth = ttk.LabelFrame(frm, text="FBA箱唛下载（可选）", padding=10)
        auth.grid(row=r, column=0, columnspan=3, sticky="we", pady=(12, 0))
        auth.columnconfigure(1, weight=1)
        auth.columnconfigure(3, weight=0)
        auth.columnconfigure(4, weight=0)

        ttk.Checkbutton(auth, text="拆分完成后自动下载 FBA 箱唛（查询→打印→下载ZIP）", variable=self.enable_fba_label_var).grid(row=0, column=0, columnspan=4, sticky="w")

        ttk.Label(auth, text="x-auth-token").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(auth, textvariable=self.token_var).grid(row=1, column=1, columnspan=3, sticky="we", padx=8, pady=(8, 0))

        ttk.Label(auth, text="cookie（至少包含 sensorsdata...）").grid(row=2, column=0, sticky="nw", pady=(8, 0))
        cookie_entry = tk.Text(auth, height=3)
        cookie_entry.grid(row=2, column=1, columnspan=3, sticky="we", padx=8, pady=(8, 0))
        self._cookie_text = cookie_entry

        ttk.Label(auth, text="账号").grid(row=3, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(auth, textvariable=self.acc_var, width=24).grid(row=3, column=1, sticky="w", padx=8, pady=(8, 0))

        ttk.Label(auth, text="密码").grid(row=3, column=2, sticky="w", pady=(8, 0))
        ttk.Entry(auth, textvariable=self.pwd_var, width=24, show="*").grid(row=3, column=3, sticky="w", padx=8, pady=(8, 0))

        ttk.Label(auth, text="打印间隔(秒)").grid(row=4, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(auth, textvariable=self.fba_cooldown_var, width=8).grid(row=4, column=1, sticky="w", padx=8, pady=(8, 0))
        ttk.Label(auth, text="（默认35；两次批量打印之间等待）").grid(row=4, column=2, columnspan=2, sticky="w", pady=(8, 0))

        ttk.Button(auth, text="自动登录获取token/cookie（可选）", command=self._auto_login).grid(row=1, column=4, rowspan=2, sticky="ns", padx=10, pady=(8, 0))

        r += 1
        btns = ttk.Frame(frm)
        btns.grid(row=r, column=0, columnspan=3, sticky="we", pady=(12, 0))
        self.run_btn = ttk.Button(btns, text="开始拆分", command=self._start)
        self.run_btn.pack(side="left")
        ttk.Button(btns, text="打开输出目录", command=self._open_outdir).pack(side="left", padx=10)

        r += 1
        prog = ttk.Frame(frm)
        prog.grid(row=r, column=0, columnspan=3, sticky="we", pady=(12, 0))
        prog.columnconfigure(0, weight=1)

        self.progress = ttk.Progressbar(prog, mode="determinate")
        self.progress.grid(row=0, column=0, sticky="we")
        self.progress_label = ttk.Label(prog, text="0%")
        self.progress_label.grid(row=0, column=1, sticky="e", padx=(8, 0))

        r += 1
        ttk.Label(frm, text="日志").grid(row=r, column=0, sticky="w", pady=(12, 0))
        r += 1
        self.log = ScrolledText(frm, height=18)
        self.log.grid(row=r, column=0, columnspan=3, sticky="nsew", pady=(6, 0))

    
    def _auto_login(self):
        try:
            acc = self.acc_var.get().strip()
            pwd = self.pwd_var.get().strip()
            if not acc or not pwd:
                messagebox.showinfo("提示", "请先输入账号和密码。")
                return
            self._append_log("🌐 开始自动登录获取 token/cookie ...")
            token, cookie = auto_login_get_token_cookie(acc, pwd, log_cb=self._append_log)
            if token:
                self.token_var.set(token)
            if hasattr(self, "_cookie_text"):
                self._cookie_text.delete("1.0", "end")
                self._cookie_text.insert("1.0", cookie)
            self._append_log("✅ 已获取 cookie（token 若为空可手动粘贴 x-auth-token）。")
            self._save_config()
        except Exception as ex:
            self._append_log(f"❌ 自动登录失败：{ex}")
            messagebox.showerror("自动登录失败", str(ex))

    def _pick_file1(self):
        p = filedialog.askopenfilename(title="选择文件1", filetypes=[("Excel", "*.xlsx;*.xls"), ("All", "*.*")])
        if p:
            self.file1_var.set(p)

    def _pick_template(self):
        p = filedialog.askopenfilename(title="选择模板文件（含公式）", filetypes=[("Excel", "*.xlsx;*.xls"), ("All", "*.*")])
        if p:
            self.template_var.set(p)

    def _pick_cfg(self):
        p = filedialog.askopenfilename(title="选择配置文件", filetypes=[("Excel", "*.xlsx;*.xls"), ("All", "*.*")])
        if p:
            self.cfg_var.set(p)

    def _pick_outdir(self):
        p = filedialog.askdirectory(title="选择输出根目录")
        if p:
            self.outdir_var.set(p)

    def _open_outdir(self):
        p = self.outdir_var.get().strip()
        if not p:
            messagebox.showinfo("提示", "请先选择输出根目录。")
            return
        if not os.path.isdir(p):
            messagebox.showerror("错误", "输出根目录不存在。")
            return
        try:
            os.startfile(p)
        except Exception:
            messagebox.showinfo("提示", f"输出目录：{p}")

    def _append_log(self, msg: str):
        self.log.insert("end", msg + "\n")
        self.log.see("end")

    def _set_progress(self, done: int, total: int, supplier_short: str):
        pct = int(done * 100 / max(total, 1))
        self.progress["maximum"] = total
        self.progress["value"] = done
        self.progress_label.config(text=f"{pct}%  ({done}/{total})  {supplier_short}")

    def _validate_inputs(self) -> Tuple[str, str, str, str, str, str, str, str, bool]:
        file1 = self.file1_var.get().strip()
        template = self.template_var.get().strip()  # optional
        cfg = self.cfg_var.get().strip()
        outdir = self.outdir_var.get().strip()
        date_str = self.date_var.get().strip()
        time_tag = self.time_var.get().strip()
        product_tag = self.product_var.get().strip()
        name = self.name_var.get().strip()
        split_supplier = bool(self.split_var.get())

        if not file1 or not os.path.isfile(file1):
            raise ValueError("请选择正确的文件1路径。")
        if not cfg or not os.path.isfile(cfg):
            raise ValueError("请选择正确的配置文件路径。")
        if not outdir:
            raise ValueError("请选择输出根目录。")
        os.makedirs(outdir, exist_ok=True)
        if not date_str:
            raise ValueError("请输入预计提货日期。")
        if not name:
            raise ValueError("请输入姓名（用于文件名前缀）。")

        # 校验日期
        parse_date(date_str)

        # 校验模板：允许为空（自动找同目录默认模板）
        _ = resolve_template_path(template)

        return file1, template, cfg, outdir, date_str, time_tag, product_tag, name, split_supplier

    def _start(self):
        try:
            file1, template, cfg, outdir, date_str, time_tag, product_tag, name, split_supplier = self._validate_inputs()
        except Exception as e:
            messagebox.showerror("输入有误", str(e))
            return

        self.run_btn.config(state="disabled")
        self.progress["value"] = 0
        self.progress_label.config(text="0%")
        self.log.delete("1.0", "end")
        self._append_log("开始处理…")

        self._save_config()

        def progress_cb(done, total, supplier_short):
            self.after(0, lambda: self._set_progress(done, total, supplier_short))

        def log_cb(msg):
            self.after(0, lambda: self._append_log(msg))

        def worker():
            try:
                outs = process_file(
                    file1=file1,
                    template_path_input=template,
                    cfg_path=cfg,
                    out_root=outdir,
                    pickup_date=date_str,
                    time_tag=time_tag,
                    product_tag=product_tag,
                    filename_name=name,
                    split_supplier_folder=split_supplier,
                    progress_cb=progress_cb,
                    log_cb=log_cb,
                )
                if not outs:
                    self.after(0, lambda: messagebox.showinfo("完成", "未找到“工厂直发”的数据行（没有输出文件）。"))
                else:
                    pickup_cell, pickup_fname = parse_date(date_str)
                    out_base = os.path.join(outdir, f"直发{pickup_fname[5:7]}{pickup_fname[8:10]}")
                    # --- 拆分完成后：可选自动下载 FBA 箱唛（不影响拆分结果） ---
                    try:
                        if bool(self.enable_fba_label_var.get()):
                            token = self.token_var.get().strip()
                            cookie = self._cookie_text.get("1.0", "end").strip() if hasattr(self, "_cookie_text") else ""
                            try:
                                cooldown_sec = int(float(self.fba_cooldown_var.get().strip() or FBA_PRINT_COOLDOWN_DEFAULT_SEC))
                            except Exception:
                                cooldown_sec = FBA_PRINT_COOLDOWN_DEFAULT_SEC
                            if not token or not cookie:
                                log_cb("⚠️ 未填写 token/cookie，跳过 FBA 箱唛下载。")
                            else:
                                for fp in outs:
                                    try:
                                        fba_download_labels_for_file(fp, token=token, cookie=cookie, log_cb=log_cb, cooldown_sec=cooldown_sec)
                                    except Exception as _ex:
                                        log_cb(f"⚠️ 箱唛下载失败（{os.path.basename(fp)}）：{_ex}")
                    except Exception as _ex2:
                        log_cb(f"⚠️ 箱唛模块异常：{_ex2}")
                    self.after(0, lambda: messagebox.showinfo("完成", f"已生成 {len(outs)} 份文件。\n输出目录：{out_base}"))
            except Exception as ex:
                # 关键修复：不要在 lambda 里直接用 ex（Python3会清掉 except 变量）
                msg = str(ex)
                self.after(0, lambda m=msg: messagebox.showerror("处理失败", m))
            finally:
                self.after(0, lambda: self.run_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _load_config(self):
        try:
            if os.path.isfile(CONFIG_PATH):
                with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                self.file1_var.set(cfg.get("file1", ""))
                self.template_var.set(cfg.get("template", ""))
                self.cfg_var.set(cfg.get("cfgfile", ""))
                self.outdir_var.set(cfg.get("outdir", ""))
                self.date_var.set(cfg.get("date", ""))
                self.time_var.set(cfg.get("time_tag", ""))
                self.product_var.set(cfg.get("product_tag", ""))
                self.name_var.set(cfg.get("name", ""))
                self.split_var.set(bool(cfg.get("split_supplier_folder", True)))
                self.token_var.set(cfg.get("x_auth_token", ""))
                if hasattr(self, "_cookie_text"):
                    self._cookie_text.delete("1.0", "end")
                    self._cookie_text.insert("1.0", cfg.get("cookie", ""))
                self.acc_var.set(cfg.get("account", ""))
                self.pwd_var.set(cfg.get("password", ""))
                self.enable_fba_label_var.set(bool(cfg.get("enable_fba_label", True)))
                self.fba_cooldown_var.set(str(cfg.get("fba_cooldown_sec", FBA_PRINT_COOLDOWN_DEFAULT_SEC)))
            else:
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                self.outdir_var.set(desktop if os.path.isdir(desktop) else os.path.expanduser("~"))
        except Exception:
            pass

    def _save_config(self):
        try:
            cfg = {
                "file1": self.file1_var.get().strip(),
                "template": self.template_var.get().strip(),
                "cfgfile": self.cfg_var.get().strip(),
                "outdir": self.outdir_var.get().strip(),
                "date": self.date_var.get().strip(),
                "time_tag": self.time_var.get().strip(),
                "product_tag": self.product_var.get().strip(),
                "name": self.name_var.get().strip(),
                "split_supplier_folder": bool(self.split_var.get()),
                "x_auth_token": self.token_var.get().strip(),
                "cookie": self._cookie_text.get("1.0", "end").strip() if hasattr(self, "_cookie_text") else "",
                "account": self.acc_var.get().strip(),
                "password": self.pwd_var.get().strip(),
                "enable_fba_label": bool(self.enable_fba_label_var.get()),
                "fba_cooldown_sec": self.fba_cooldown_var.get().strip(),
            }
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _on_close(self):
        self._save_config()
        self.destroy()


if __name__ == "__main__":
    App().mainloop()